#!/usr/bin/env python3
"""
数据清洗与 Excel 导出
读取 raw_data.json，去重、格式统一、生成带样式的 .xlsx 文件。

用法:
  python clean_and_export.py --input raw_data.json --output jobs_clean.xlsx
"""

import json
import re
import sys
import os
import argparse
import tempfile
from datetime import datetime

from utils.protocol import emit, emit_fatal

try:
    import pandas as pd
except ImportError:
    emit_fatal("DEP_PANDAS", "缺少 pandas 库，请执行: pip install pandas")

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    emit_fatal("DEP_OPENPYXL", "缺少 openpyxl 库，请执行: pip install openpyxl")


def clean_data(raw_path, output_path, min_count=25):
    """主清洗流程，返回 (success, clean_count)"""
    # 读取原始数据
    with open(raw_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    items = raw.get("list_items", [])
    details = raw.get("detail_items", [])
    keyword = raw.get("keyword", "")
    city = raw.get("city", "")

    if not items:
        emit("clean.empty", "没有数据可清洗", status="warn")
        return False, 0

    # 构建详情查找表：url → detail
    detail_map = {}
    for d in details:
        url = d.get("url", "")
        if url:
            detail_map[url] = d

    rows = []
    for item in items:
        link = item.get("link", "")
        detail = detail_map.get(link, {})

        # JD 文本：优先取详情页完整文本，取不到则从列表页拼接
        jd_text = detail.get("jd_text", "")
        if not jd_text:
            jd_text = "（未获取到详细 JD，列表页信息："
            tags = item.get("tags", [])
            if tags:
                jd_text += " | ".join(tags)
            jd_text += "）"

        rows.append({
            "岗位名称": item.get("title", ""),
            "公司名称": item.get("company", ""),
            "工作地点": item.get("location", ""),
            "JD摘要": jd_text,
            "岗位链接": link,
        })

    df = pd.DataFrame(rows)

    # 去重：同公司同岗位保留一条
    before = len(df)
    df = df.drop_duplicates(subset=["岗位名称", "公司名称"], keep="first")
    after = len(df)

    # 过滤实习岗位
    before_filter = len(df)
    df = df[~df["岗位名称"].str.contains("实习", na=False)]
    after_filter = len(df)
    if before_filter != after_filter:
        emit("clean.filter_intern", f"已过滤 {before_filter - after_filter} 个实习岗位")

    clean_count = len(df)

    # 检查是否达到最少条数要求
    if clean_count < min_count:
        emit("clean.insufficient",
             f"清洗后仅 {clean_count} 条，未达到最低要求 {min_count} 条",
             data={"clean_count": clean_count, "min_count": min_count})
        # 仍然导出已有数据
    else:
        emit("clean.count_ok",
             f"清洗后 {clean_count} 条，达到最低要求 {min_count} 条",
             data={"clean_count": clean_count, "min_count": min_count})

    # 写入 Excel（带格式）
    df.to_excel(output_path, index=False, sheet_name="AI岗位数据",
                engine="openpyxl")

    # 格式化
    wb = load_workbook(output_path)
    ws = wb.active

    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 表头样式
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据区域样式
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = cell_align
            cell.border = thin_border

    # 列宽
    col_widths = {
        "岗位名称": 22,
        "公司名称": 28,
        "工作地点": 16,
        "JD摘要": 80,
        "岗位链接": 40,
    }
    for col_idx, col_name in enumerate(df.columns, 1):
        width = col_widths.get(col_name, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{ws.max_row}"

    wb.save(output_path)

    emit("done",
         f"清洗完成：{before} 条 → 去重后 {after} 条 → 最终 {clean_count} 条 → 已导出 {output_path}",
         data={"input_rows": before, "after_dedup": after, "clean_count": clean_count, "output": output_path})
    return True, clean_count


def main():
    parser = argparse.ArgumentParser(description="数据清洗与 Excel 导出")
    default_input = os.path.join(tempfile.gettempdir(), "ai-job-hunter", "raw_data.json")
    parser.add_argument("--input", default=default_input, help="输入 raw_data.json 路径")
    parser.add_argument("--output", default="jobs_clean.xlsx", help="输出 .xlsx 路径")
    parser.add_argument("--min-count", type=int, default=25, help="最少有效数据条数，默认 25")
    args = parser.parse_args()

    success, clean_count = clean_data(args.input, args.output, min_count=args.min_count)

    if not success:
        sys.exit(1)

    if clean_count < args.min_count:
        sys.exit(2)

    # 清理中间文件
    temp_dir = os.path.join(tempfile.gettempdir(), "ai-job-hunter")
    if os.path.exists(temp_dir):
        import shutil
        try:
            shutil.rmtree(temp_dir)
        except OSError as e:
            emit("clean.cleanup", f"清理临时文件失败: {e}",
                 status="warn", warnings=[str(e)], stream=sys.stderr)


if __name__ == "__main__":
    main()
